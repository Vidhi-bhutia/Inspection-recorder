import streamlit as st
import SpeechRecognition as sr
import openpyxl
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
from gtts import gTTS
import tempfile
import pygame
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO

# Initialize the speech recognition engine
recognizer = sr.Recognizer()

# Initialize Pygame mixer for playing audio
pygame.mixer.init()

def speak(text):
    try:
        tts = gTTS(text=text, lang='en')
        temp_file_path = tempfile.mktemp(suffix='.mp3')
        tts.save(temp_file_path)
        pygame.mixer.music.load(temp_file_path)
        pygame.mixer.music.play()
        while pygame.mixer.music.get_busy():
            pygame.time.Clock().tick(10)
    except Exception as e:
        st.error(f"Error in text-to-speech: {e}")

def record_speech(prompt):
    speak(prompt)
    with sr.Microphone() as source:
        recognizer.adjust_for_ambient_noise(source, duration=1)
        st.write(f"Listening for: '{prompt}'")
        try:
            audio = recognizer.listen(source, timeout=20, phrase_time_limit=20)
            reply = recognizer.recognize_google(audio)
            st.write(f"You said: '{reply}'")
            return reply
        except sr.UnknownValueError:
            st.write(f"Sorry, I didn't catch that. Can you please repeat?")
            return "Unknown"
        except sr.RequestError:
            st.write(f"Sorry, I'm having trouble with speech recognition.")
            return "Error"
        except sr.WaitTimeoutError:
            st.write(f"Listening timed out. Please try again.")
            return "Timeout"

# Define file name
file_name = "Truck_Inspection_Records.xlsx"

# Check if file exists
if os.path.exists(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Truck Inspections"

    headers = [
        "Inspection ID", "Truck Serial Number", "Truck Model", "Inspector Name",
        "Inspection Employee ID", "Date & Time of Inspection", "Location of Inspection",
        "Geo Coordinates of Inspection", "Service Meter Hours", "Inspector Signature",
        "Customer Name / Company Name", "CAT Customer ID",
        "Tire Pressure for Left Front", "Tire Pressure for Right Front",
        "Tire Condition for Left Front", "Tire Condition for Right Front",
        "Tire Pressure for Left Rear", "Tire Pressure for Right Rear",
        "Tire Condition for Left Rear", "Tire Condition for Right Rear",
        "Overall Tire Summary", "Tire Images",
        "Battery Make", "Battery replacement date", "Battery Voltage",
        "Battery Water level", "Condition of Battery", "Any Leak / Rust in battery",
        "Battery overall Summary", "Battery Images",
        "Exterior Rust/Dent/Damage", "Oil leak in Suspension", "Exterior Summary", "Exterior Images",
        "Brake Fluid level", "Brake Condition for Front", "Brake Condition for Rear",
        "Emergency Brake", "Brake Overall Summary", "Brake Images",
        "Engine Rust/Dents/Damage", "Engine Oil Condition", "Engine Oil Color",
        "Brake Fluid Condition", "Brake Fluid Color", "Any oil leak in Engine",
        "Engine Overall Summary", "Engine Images",
        "Voice of Customer", "Customer Feedback Images"
    ]

    ws.append(headers)

# Streamlit app UI
st.title("Truck Inspection Data Collection")

# Automatically generate an ID
inspection_data = []
inspection_data.append(str(len(ws['A']) + 1))  # Inspection ID is auto-incremented

# Collecting inputs for each component
questions = {
    "Truck Serial Number": "Please provide the Truck Serial Number.",
    "Truck Model": "Please provide the Truck Model.",
    "Inspector Name": "Please provide your name.",
    "Inspection Employee ID": "Please provide your Employee ID.",
    "Location of Inspection": "Please provide the Location of Inspection.",
    "Geo Coordinates of Inspection": "Please provide the Geo Coordinates of Inspection if available.",
    "Service Meter Hours": "Please provide the Service Meter Hours.",
    "Inspector Signature": "Please provide your Signature verbally.",
    "Customer Name / Company Name": "Please provide the Customer Name or Company Name.",
    "CAT Customer ID": "Please provide the CAT Customer ID.",
    "Tire Pressure for Left Front": "Please provide the Tire Pressure for Left Front.",
    "Tire Pressure for Right Front": "Please provide the Tire Pressure for Right Front.",
    "Tire Condition for Left Front": "Please describe the Tire Condition for Left Front.",
    "Tire Condition for Right Front": "Please describe the Tire Condition for Right Front.",
    "Tire Pressure for Left Rear": "Please provide the Tire Pressure for Left Rear.",
    "Tire Pressure for Right Rear": "Please provide the Tire Pressure for Right Rear.",
    "Tire Condition for Left Rear": "Please describe the Tire Condition for Left Rear.",
    "Tire Condition for Right Rear": "Please describe the Tire Condition for Right Rear.",
    "Overall Tire Summary": "Please provide an Overall Tire Summary.",
    "Battery Make": "Please provide the Battery Make.",
    "Battery replacement date": "Please provide the Battery replacement date.",
    "Battery Voltage": "Please provide the Battery Voltage.",
    "Battery Water level": "Please describe the Battery Water level.",
    "Condition of Battery": "Please describe the Condition of Battery.",
    "Any Leak / Rust in battery": "Is there any Leak or Rust in the battery? (Yes/No)",
    "Battery overall Summary": "Please provide an Overall Battery Summary.",
    "Exterior Rust/Dent/Damage": "Is there any Rust, Dent or Damage to the Exterior? (Yes/No)",
    "Oil leak in Suspension": "Is there any Oil leak in the Suspension? (Yes/No)",
    "Exterior Summary": "Please provide an Overall Exterior Summary.",
    "Brake Fluid level": "Please describe the Brake Fluid level.",
    "Brake Condition for Front": "Please describe the Brake Condition for Front.",
    "Brake Condition for Rear": "Please describe the Brake Condition for Rear.",
    "Emergency Brake": "Please describe the Emergency Brake condition.",
    "Brake Overall Summary": "Please provide an Overall Brake Summary.",
    "Engine Rust/Dents/Damage": "Is there any Rust, Dents or Damage in the Engine? (Yes/No)",
    "Engine Oil Condition": "Please describe the Engine Oil Condition.",
    "Engine Oil Color": "Please describe the Engine Oil Color.",
    "Brake Fluid Condition": "Please describe the Brake Fluid Condition.",
    "Brake Fluid Color": "Please describe the Brake Fluid Color.",
    "Any oil leak in Engine": "Is there any oil leak in the Engine? (Yes/No)",
    "Engine Overall Summary": "Please provide an Overall Engine Summary.",
    "Voice of Customer": "Please provide any feedback from the Customer."
}

# Insert the date and time at the correct position
date_time_of_inspection = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
inspection_data.insert(5, date_time_of_inspection)

# Ask questions and collect responses
for key, question in questions.items():
    while True:
        response = record_speech(question)
        if response not in ["Unknown", "Error", "Timeout"]:
            inspection_data.append(response)
            break
        else:
            speak(f"Sorry, I couldn't get you. Can you please repeat the {key.lower()}?")

# Save data to Excel and generate PDF when the form is submitted
if st.button("Save Inspection Data"):
    ws.append(inspection_data)
    wb.save(file_name)
    st.success(f"Inspection record saved to {file_name}.")

    # Create a PDF and provide download link
    pdf_buffer = BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=letter)
    width, height = letter
    y_position = height - 50

    # Add title
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, y_position, "Truck Inspection Report")
    y_position -= 30

    # Add inspection data
    c.setFont("Helvetica", 12)
    for header, data in zip(headers, inspection_data):
        c.drawString(50, y_position, f"{header}: {data}")
        y_position -= 20

    c.save()

    pdf_buffer.seek(0)
    st.download_button(
        label="Download PDF",
        data=pdf_buffer,
        file_name="Truck_Inspection_Report.pdf",
        mime="application/pdf"
    )
